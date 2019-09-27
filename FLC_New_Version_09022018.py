#!/usr/bin/env python
# coding: utf-8

# # Facility List Coder, version 12/02/2018  

# The Facility List coder (FLC) is an open source tool that allow efficiently combine GIS analysis with standard data techniques. Besides the data management tools, FLC code retrieves GIS information on facilities location using two open source datasets: Google Maps and Open Street Maps. 
# 
# FLC is built upon two main requirements. 
# - First, researchers will need to provide a specific data set for the specific location of the reference point (e.g. school, university, among others). We call them as location of interest (LI). 
# - Second in order to classify the results obtained from the spatial query on the traditional GIS engines, researchers will need to define a set of key words or metadata that allow the algorithm classify the facilities. Based on the literature (Caspi et al. 2012; Lytle and Sokol 2017; Wilkins et al. 2017), we have developed a multi-language key words list based on the European context that allows to classify each facility within a pre-defined category. These categories could be modified in order to fulfill specific needs of researchers related to geographical location, languages or research questions. This pre-defined key word offers an important innovation for the research on food community environment in the European context. The empirical studies for Europe often use categories created for United States, which might miss-estimate the particularities of European food traditions. Nonetheless, this list will be easily extended depending on the needs of the researchers. We called this dimension as key words by categories (KW).
# 
# 

# ## Code Preliminaries

# FLC code retrieves GIS information on facilities location using two open source datasets: Google Maps and Open Street Maps. All the code was developed on Python 2.7.

# ###  Defining functions

# Now, we need to create functions that will be use during the process

# In[49]:


### Function to save results 
## 
def save_result(res,prints):
    for place in res.places:
        place.get_details()
        if prints==1:
            print place.name.encode('utf-8'),place.types,place.place_id
        else:
            if place.place_id not in places_flc:
                raw={'place_name':place.name,'place_lat':float(place.geo_location['lat']),'place_lng':float(place.geo_location['lng']),'place_address':place.formatted_address,'google_id':place.place_id,'place_types':place.types,'place_web':place.url}
                places_flc[place.place_id]=raw
                    

## Read 
## This functions was made to test the results
def read_result(res):
    for place in res.places:
        place.get_details()
        print place.name.encode('utf-8'),place.types,place.place_id
    
                
### Save Results
## Save Results
def sav_final(place,cat,final_ds):
    ## Save Final Data Set
    raw=places_flc_cleaned[place]
    raw['categ']=cat
    final_ds.append(raw)
    ## Delete the dictionary


# ### Loading packages 

# Here we need to load the needed packages

# In[50]:


# coding=utf-8
###########################
###### Libraries and Initial SetUp
###########################

### My Data key
YOUR_API_KEY = 'YOUR_API_KEY'

## --- Google maps
import googlemaps
gmaps = googlemaps.Client(key=YOUR_API_KEY)
from datetime import datetime

# ----- Google Place
from googleplaces import GooglePlaces, types, lang
import google_streetview.api
google_places = GooglePlaces(YOUR_API_KEY)

# ----- Download the data
import urllib2 # Download the files
import os, sys ## create a new 
import pandas as pd
from xlsxwriter.utility import xl_rowcol_to_cell

# ------ from openpyxl import load_workbook
from openpyxl import load_workbook
import fiona

# ------ Import Regex
import re

# ----- Load information
import json
 


# In[ ]:





# ## Input 1:  Key Words

# In this section we load the key words that will be used for the classification of the results.

# ### Key words for the Catalan context

# The first part of the key words are a detailed description of the economic activities in a specific region. In the case of Catalunya, we employed: "Criteris registrals pe a establiments minoristes d'alimentació en Catalunya" from the Regional Government. This document brings a details description of the different types of establishments present in the region. 
# 
# All the keywords were translated in English, Spanish and Catalan. Everything is save at: **Palabras_clave_v0_120917.xlsx**

# In[51]:


data_keywords="/Users/juancarlosmunoz/Dropbox/Obsogenic_Project/01_DataBase_Georeferencion/Palabras_clave_v0_120917.xlsx"
wb = load_workbook(filename = data_keywords)
palabras = wb['Plabras_Claves_Google']

### Create the dictionary
categ_est={}
keywords=[]
for row in range(2,54):
    cat=str(palabras.cell(row=row, column=1).value.encode('utf-8')).lower().strip()
    try:
        categ_est[cat]=[]
    except:
        pass

#### Now the key words
for row in range(2,54):
    cat=str(palabras.cell(row=row, column=1).value.encode('utf-8')).lower().strip()
    for col in [2,3,4]:
        if palabras.cell(row=row, column=col).value is not None:
            word=str(palabras.cell(row=row, column=col).value.encode('utf-8')).lower().strip()
            if word not in categ_est[cat]:
                categ_est[cat].append(str(word))


# ### Type Excluded

# In order to optimize the query over the google maps we restrict the search to particular set of places types ([here the list of accepted types in Google](https://developers.google.com/places/supported_types)).
# 
# In order to exclude/include a given category you only have to change: **google_types.xlsx**

# In[52]:


data_types="/Users/juancarlosmunoz/Dropbox/Obsogenic_Project/01_DataBase_Georeferencion/google_types.xlsx"
wb = load_workbook(filename = data_types)
palabras = wb['Sheet1']
types_accepted=[]
types_excluded=[]
for row in range(2,98):
    selected=int(palabras.cell(row=row, column=2).value)
    if selected==0:
        types_accepted.append(str(palabras.cell(row=row, column=1).value))
    if selected==1:
        types_excluded.append(str(palabras.cell(row=row, column=1).value))

print types_accepted


# ## Input 2: Defining Location  of interest (LI)

# In this section we load the location of the point of interest. For the validation procedure, we include the Grids. 
# 
# Note that you need a shapefile with the centroids. It can be generalized even more using any type of argument.

# In[53]:


grids="/Users/juancarlosmunoz/Dropbox/Familia_Munoz_Arcila/05_Paper/shp/final_sample_15122017_centroids_WSG84.shp"
grids_data={}
with fiona.open(grids,'r') as shp:
    for feat in shp:
        grid_id=feat['properties']['grid_id']
        lng,lat=feat['geometry']['coordinates']
        grids_data[grid_id]={'lat':lat,'lng':lng}


# ## Facility List Coder in action: Building the dataset

# In this section we run the spatial query in Google maps based on the point of interest. Technically, the FLC will get all the location within an specific folder, then we will classify then using the key words.
# 
# The first step is defining the data set.

# In[54]:


# Data Information
places_flc={}


# ### Spatial Query Strategy 1: Using only Google types

# This strategy will get all the places within a buffer that belongs to a specific type (look above to check the list o types

# ### Spatial Query Strategy 2: Using only Google types

# In[55]:


i=1

### To Print results
prints=0

#for grid in [755]:
for grid in grids_data:
    print "%d of 301 (grid=%d)" %(i,grid)
    i+=1
    ### First Step Search
    for typ in types_accepted:
   #for typ in ['restaurant']:
        ### First 20 results
        res1=google_places.nearby_search(lat_lng=grids_data[grid],radius=100,types=[typ])
        save_result(res1,prints)
        # For more than 20 results
        try:
            res2 = google_places.nearby_search(pagetoken=res1.next_page_token)
            save_result(res2,prints)
            try:
                res3 = google_places.nearby_search(pagetoken=res2.next_page_token)
                save_result(res3,prints)
                try:
                    res4 = google_places.nearby_search(pagetoken=res3.next_page_token)
                    save_result(res4,prints)
                except:
                    pass  
            except:
                pass    
        except:
            pass


# This data is complementary to the last one, but it will search by keywords instead of type. This part will search at any information gathered by google.

# In[59]:


#types=types_accepted+['supermarket']

### To Print results
prints=0

i=1
for grid in grids_data:
    print "%d of 301 (grid=%d)" %(i,grid)
    i+=1
#for grid in [1020]:
    ### First Step Search
    for typ in types_accepted:
        #print typ
        ### First 20 results
        res1=google_places.nearby_search(lat_lng=grids_data[grid],radius=100,keyword=typ)
        # For more than 20 results
        save_result(res1,prints)
        try:
            res2 = google_places.nearby_search(pagetoken=res1.next_page_token)
            save_result(res2,prints)
            try:
                res3 = google_places.nearby_search(pagetoken=res2.next_page_token)
                save_result(res3,prints)
                try:
                    res4 = google_places.nearby_search(pagetoken=res3.next_page_token)
                    save_result(res4,prints)
                except:
                    pass    
            except:
                pass
        except:
            pass


# In[62]:


print len(places_flc)


# ### Save RawData

# Before classification, we save the entire data set.

# In[63]:


import json
### Save
json = json.dumps(places_flc)
f = open("/Users/juancarlosmunoz/Dropbox/Obsogenic_Project/02_Output_DataBase/flc_rawresults_grids_13022018.json","w")
f.write(json)
f.close()

## Open
#file=open("/Users/juancarlosmunoz/Dropbox/Obsogenic_Project/flc_rawresults_grids.json", 'r')
#data = json.load(places_flc)


# ## Facility List Coder in action: Classifying the places

# Now, once all the places are gather using the two strategies, we now need to clean and classify then.

# ### Clean Data Set

# First we delete those establishments that will be excluded.

# In[81]:


### First data
places_flc_cleaned={}

# Excluded Types
excluded_types=['health','finance','pharmacy','electrician','church','parking']
# Excluded Names (part of)
excluded_names=['parking']

#### First Step ---> Those in the excluded categories
for i in places_flc:
    ### Check for those with a excluded
    name=places_flc[i]['place_name'].lower().encode('utf-8').strip().split()
    if any(x in places_flc[i]['place_types'] for x in types_excluded):
        pass
    if any(x in name for x in excluded_names):
        pass
    else:
        places_flc_cleaned[i]=places_flc[i]


# In[82]:


final_ds=[]
del_id=[]
## Step 1: Classification based on type of place
for place in places_flc_cleaned:
    ## Step 0 - Using Type
    if 'bakery' in places_flc_cleaned[place]['place_types']:
        sav_final(place,'harinas y derivados',final_ds)
        del_id.append(str(place))
        
    if any(x in places_flc_cleaned[place]['place_types'] for x in [u'grocery_or_supermarket',u'convenience_store']):
        sav_final(place,'establecimientos polivalentes',final_ds)
        del_id.append(str(place))
        
    if any(x in places_flc_cleaned[place]['place_types'] for x in [u'cafe',u'restaurant',u'meal_delivery',u'meal_takeaway',u'bar']):
        sav_final(place,'comidas preparadas',final_ds)
        del_id.append(str(place))


# In[83]:


## Step 2: Classification based on names
for place in places_flc_cleaned:
    if place not in del_id:
        for cat in categ_est:
            for word in categ_est[cat]:
                if word!='bar':
                    w_test=str(places_flc_cleaned[place]['place_name'].encode('utf-8')).lower()
                    q_test="^.*"+word+".*$"
                    query=re.search(q_test,w_test)
                    if query is not None:
                        sav_final(place,cat,final_ds)
                        del_id.append(str(place))
                        print places_flc_cleaned[place]['place_name']

                #print query
                #try:
                #    print "----------"
                #    print len(query.group(0))
                #    print "----------"
                #except:
                #    pass
                #if query.group(0) is not NULL:
                #    print "HELLO" 
                # =str(palabras.cell(row=row, column=col).value).lower().strip().decode('ascii')
            
            
        


# In[84]:


for place in places_flc_cleaned:
    if place not in del_id:
        if 'food' in places_flc_cleaned[place]['place_types']:
            pass
        else:
            print places_flc_cleaned[place]['place_name'],places_flc_cleaned[place]['place_types']


# ## Save Data Set
# 
# 

# In[85]:


#### Final create the dataset
file_name="/Users/juancarlosmunoz/Dropbox/Obsogenic_Project/02_Output_DataBase/DataBase_v2_Feb2018_grids.xlsx"
df = pd.DataFrame(final_ds)
writer_orig = pd.ExcelWriter(file_name, engine='xlsxwriter')
df.to_excel(writer_orig, index=False, sheet_name='report')
writer_orig.save()


# ### Extra Checks (no need to run)

# In[70]:


## Check whether especific store was saved
df.loc[df['google_id'] == 'ChIJAQCEozC1pBIRzf1jRrgKvcQ']


# In[12]:






for typ in types_accepted:
    res1=google_places.nearby_search(lat_lng=grids_data[1175],radius=50,type='bakery')
    # For more than 20 results
    print "---- %s -----" %(typ)
    read_result(res1)
    try:
        res2 = google_places.nearby_search(pagetoken=res1.next_page_token)
        print res2
        read_result(res2)
        try:
            res3 = google_places.nearby_search(pagetoken=res2.next_page_token)
            print res3
            read_result(res3)
            try:
                res4 = google_places.nearby_search(pagetoken=res3.next_page_token)
                print res4
                read_result(res4)
            except:
                pass    
        except:
            pass   
    except:
        pass

            


# In[409]:



print types_accepted


# In[ ]:


ChIJAQCEozC1pBIRzf1jRrgKvcQ

